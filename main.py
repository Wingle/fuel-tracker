from collections import OrderedDict

import csv
import io
import math
import secrets
from datetime import date, datetime
from typing import Optional

import bcrypt
import openpyxl
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, model_validator
from sqlalchemy import asc, inspect
from sqlalchemy.orm import Session

from database import Base, engine, get_db
from models import FuelRecord, User, UserSession, Vehicle

# ---------------------------------------------------------------------------
# Create tables
# ---------------------------------------------------------------------------
Base.metadata.create_all(bind=engine)

# ---------------------------------------------------------------------------
# Migrate legacy data
# ---------------------------------------------------------------------------
def _migrate_legacy_records():
    insp = inspect(engine)
    cols = [c["name"] for c in insp.get_columns("fuel_records")]
    if "vehicle_id" not in cols:
        return
    from database import SessionLocal
    db = SessionLocal()
    try:
        orphans = db.query(FuelRecord).filter(FuelRecord.vehicle_id.is_(None)).count()
        if orphans == 0:
            return
        default = db.query(Vehicle).filter(Vehicle.name == "默认车辆").first()
        if not default:
            default = Vehicle(name="默认车辆", plate_number="")
            db.add(default)
            db.flush()
        db.query(FuelRecord).filter(FuelRecord.vehicle_id.is_(None)).update(
            {FuelRecord.vehicle_id: default.id}
        )
        db.commit()
        print(f"[migrate] {orphans} 条旧记录已迁移到「默认车辆」(id={default.id})")
    finally:
        db.close()

_migrate_legacy_records()

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = FastAPI(title="油耗记录工具")


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token() -> str:
    return secrets.token_hex(32)

def get_current_user(
    authorization: Optional[str] = Header(None),
    token: Optional[str] = Query(None, alias="token"),
    db: Session = Depends(get_db),
) -> User:
    """Extract user from Bearer header or ?token= query param (for CSV export)."""
    auth_token = None
    if authorization and authorization.startswith("Bearer "):
        auth_token = authorization[7:]
    elif token:
        auth_token = token
    if not auth_token:
        raise HTTPException(status_code=401, detail="未登录")
    session = db.query(UserSession).filter(UserSession.token == auth_token).first()
    if not session:
        raise HTTPException(status_code=401, detail="登录已过期，请重新登录")
    return session.user


# ---------------------------------------------------------------------------
# Auth schemas
# ---------------------------------------------------------------------------
class AuthRequest(BaseModel):
    username: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    security_question: str
    security_answer: str

class AuthResponse(BaseModel):
    token: str
    user_id: int
    username: str

class ForgotPasswordQuestionRequest(BaseModel):
    username: str

class ForgotPasswordQuestionResponse(BaseModel):
    security_question: str

class ForgotPasswordResetRequest(BaseModel):
    username: str
    security_answer: str
    new_password: str


# ---------------------------------------------------------------------------
# Auth API (no token required)
# ---------------------------------------------------------------------------
@app.post("/api/auth/register", response_model=AuthResponse)
def register(payload: RegisterRequest, db: Session = Depends(get_db)):
    username = payload.username.strip()
    if len(username) < 2 or len(username) > 50:
        raise HTTPException(status_code=400, detail="用户名长度需要 2-50 个字符")
    if len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="密码长度至少 6 个字符")
    if not payload.security_question.strip():
        raise HTTPException(status_code=400, detail="请输入安全提示问题")
    if not payload.security_answer.strip():
        raise HTTPException(status_code=400, detail="请输入安全提示问题的答案")
    existing = db.query(User).filter(User.username == username).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户名已被注册")
    user = User(
        username=username,
        password_hash=hash_password(payload.password),
        security_question=payload.security_question.strip(),
        security_answer_hash=hash_password(payload.security_answer.strip()),
    )
    db.add(user)
    db.flush()
    token = create_token()
    db.add(UserSession(user_id=user.id, token=token))
    db.commit()
    return AuthResponse(token=token, user_id=user.id, username=user.username)


@app.post("/api/auth/login", response_model=AuthResponse)
def login(payload: AuthRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == payload.username.strip()).first()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_token()
    db.add(UserSession(user_id=user.id, token=token))
    db.commit()
    return AuthResponse(token=token, user_id=user.id, username=user.username)


@app.post("/api/auth/logout")
def logout(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    if authorization and authorization.startswith("Bearer "):
        token = authorization[7:]
        session = db.query(UserSession).filter(UserSession.token == token).first()
        if session:
            db.delete(session)
            db.commit()
    return {"ok": True}


@app.get("/api/auth/me")
def get_me(user: User = Depends(get_current_user)):
    return {"user_id": user.id, "username": user.username}


# ---------------------------------------------------------------------------
# Forgot password API (no token required)
# ---------------------------------------------------------------------------
@app.post("/api/auth/forgot-password/question", response_model=ForgotPasswordQuestionResponse)
def forgot_password_question(payload: ForgotPasswordQuestionRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == payload.username.strip()).first()
    if not user or not user.security_question:
        raise HTTPException(status_code=404, detail="用户不存在或未设置安全提示问题")
    return ForgotPasswordQuestionResponse(security_question=user.security_question)


@app.post("/api/auth/forgot-password/reset")
def forgot_password_reset(payload: ForgotPasswordResetRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == payload.username.strip()).first()
    if not user or not user.security_answer_hash:
        raise HTTPException(status_code=404, detail="用户不存在")
    if not verify_password(payload.security_answer.strip(), user.security_answer_hash):
        raise HTTPException(status_code=400, detail="安全提示问题的答案不正确")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="新密码长度至少 6 个字符")
    user.password_hash = hash_password(payload.new_password)
    # 清除所有现有会话，强制重新登录
    db.query(UserSession).filter(UserSession.user_id == user.id).delete()
    db.commit()
    return {"ok": True, "message": "密码重置成功，请重新登录"}


# ---------------------------------------------------------------------------
# Pydantic schemas — Vehicle
# ---------------------------------------------------------------------------
class VehicleCreate(BaseModel):
    name: str
    plate_number: Optional[str] = ""

class VehicleUpdate(VehicleCreate):
    pass

class VehicleOut(BaseModel):
    id: int
    name: str
    plate_number: Optional[str]
    created_at: datetime
    model_config = {"from_attributes": True}


# ---------------------------------------------------------------------------
# Pydantic schemas — FuelRecord
# ---------------------------------------------------------------------------
class RecordCreate(BaseModel):
    vehicle_id: int
    date: date
    mileage: float
    volume: Optional[float] = None
    unit_price: Optional[float] = None
    total_price: Optional[float] = None
    fuel_type: Optional[str] = "92#"
    note: Optional[str] = ""

    @model_validator(mode="after")
    def auto_calculate(self):
        v, u, t = self.volume, self.unit_price, self.total_price
        count = sum(x is not None and x > 0 for x in [v, u, t])
        if count >= 2:
            if v and u and not t:
                self.total_price = round(v * u, 2)
            elif v and t and not u:
                self.unit_price = round(t / v, 2) if v else None
            elif u and t and not v:
                self.volume = round(t / u, 2) if u else None
        return self

class RecordUpdate(RecordCreate):
    pass

class RecordOut(BaseModel):
    id: int
    vehicle_id: Optional[int]
    date: date
    mileage: float
    volume: Optional[float]
    unit_price: Optional[float]
    total_price: Optional[float]
    fuel_type: Optional[str]
    note: Optional[str]
    created_at: datetime
    distance: Optional[float] = None
    fuel_consumption: Optional[float] = None
    cost_per_km: Optional[float] = None
    model_config = {"from_attributes": True}

class PaginatedRecords(BaseModel):
    items: list[RecordOut]
    total: int
    page: int
    page_size: int
    total_pages: int

class StatsOut(BaseModel):
    total_mileage: float
    total_cost: float
    total_volume: float
    avg_consumption: Optional[float]
    daily_mileage: Optional[float]
    record_count: int

class PeriodSummaryItem(BaseModel):
    period: str
    mileage: float
    volume: float
    cost: float
    avg_consumption: Optional[float]
    daily_mileage: Optional[float]
    record_count: int

class PeriodSummaryOut(BaseModel):
    mode: str
    items: list[PeriodSummaryItem]

class ImportResult(BaseModel):
    imported: int
    skipped: int
    errors: list[str]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _get_user_vehicle(db: Session, user: User, vehicle_id: int) -> Vehicle:
    """Fetch a vehicle owned by the user, or 404."""
    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.user_id == user.id).first()
    if not v:
        raise HTTPException(status_code=404, detail="车辆不存在")
    return v

def _query_records(db: Session, vehicle_id: int):
    return (
        db.query(FuelRecord)
        .filter(FuelRecord.vehicle_id == vehicle_id)
        .order_by(asc(FuelRecord.date), asc(FuelRecord.mileage))
        .all()
    )

def enrich_records(records: list[FuelRecord]) -> list[dict]:
    sorted_recs = sorted(records, key=lambda r: (r.date, r.mileage))
    result = []
    for i, rec in enumerate(sorted_recs):
        data = {
            "id": rec.id, "vehicle_id": rec.vehicle_id,
            "date": rec.date, "mileage": rec.mileage,
            "volume": rec.volume, "unit_price": rec.unit_price,
            "total_price": rec.total_price, "fuel_type": rec.fuel_type,
            "note": rec.note, "created_at": rec.created_at,
            "distance": None, "fuel_consumption": None, "cost_per_km": None,
        }
        if i > 0:
            prev = sorted_recs[i - 1]
            dist = rec.mileage - prev.mileage
            if dist > 0:
                data["distance"] = round(dist, 1)
                if rec.volume and rec.volume > 0:
                    data["fuel_consumption"] = round(rec.volume / dist * 100, 2)
                if rec.total_price and rec.total_price > 0:
                    data["cost_per_km"] = round(rec.total_price / dist, 2)
        result.append(data)
    result.reverse()
    return result


# ---------------------------------------------------------------------------
# Vehicle API (auth required)
# ---------------------------------------------------------------------------
@app.get("/api/vehicles", response_model=list[VehicleOut])
def list_vehicles(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(Vehicle).filter(Vehicle.user_id == user.id).order_by(asc(Vehicle.created_at)).all()

@app.post("/api/vehicles", response_model=VehicleOut)
def create_vehicle(payload: VehicleCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    v = Vehicle(user_id=user.id, name=payload.name, plate_number=payload.plate_number)
    db.add(v)
    db.commit()
    db.refresh(v)
    return v

@app.put("/api/vehicles/{vehicle_id}", response_model=VehicleOut)
def update_vehicle(vehicle_id: int, payload: VehicleUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    v = _get_user_vehicle(db, user, vehicle_id)
    v.name = payload.name
    v.plate_number = payload.plate_number
    db.commit()
    db.refresh(v)
    return v

@app.delete("/api/vehicles/{vehicle_id}")
def delete_vehicle(vehicle_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    v = _get_user_vehicle(db, user, vehicle_id)
    count = db.query(FuelRecord).filter(FuelRecord.vehicle_id == vehicle_id).count()
    if count > 0:
        raise HTTPException(status_code=400, detail=f"该车辆下还有 {count} 条加油记录，请先删除记录")
    db.delete(v)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Record API (auth required)
# ---------------------------------------------------------------------------
@app.get("/api/records", response_model=PaginatedRecords)
def list_records(
    vehicle_id: int = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_user_vehicle(db, user, vehicle_id)  # ownership check
    all_records = _query_records(db, vehicle_id)
    enriched = enrich_records(all_records)
    total = len(enriched)
    total_pages = max(1, math.ceil(total / page_size))
    page = min(page, total_pages)
    start = (page - 1) * page_size
    items = enriched[start:start + page_size]
    return PaginatedRecords(items=items, total=total, page=page, page_size=page_size, total_pages=total_pages)

@app.post("/api/records", response_model=RecordOut)
def create_record(payload: RecordCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _get_user_vehicle(db, user, payload.vehicle_id)
    rec = FuelRecord(
        vehicle_id=payload.vehicle_id, date=payload.date, mileage=payload.mileage,
        volume=payload.volume, unit_price=payload.unit_price, total_price=payload.total_price,
        fuel_type=payload.fuel_type, note=payload.note,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    all_records = _query_records(db, payload.vehicle_id)
    enriched = enrich_records(all_records)
    for e in enriched:
        if e["id"] == rec.id:
            return e
    return rec

@app.put("/api/records/{record_id}", response_model=RecordOut)
def update_record(record_id: int, payload: RecordUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _get_user_vehicle(db, user, payload.vehicle_id)
    rec = db.query(FuelRecord).filter(FuelRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="记录不存在")
    # Verify the record belongs to user's vehicle
    if rec.vehicle_id:
        _get_user_vehicle(db, user, rec.vehicle_id)
    rec.vehicle_id = payload.vehicle_id
    rec.date = payload.date
    rec.mileage = payload.mileage
    rec.volume = payload.volume
    rec.unit_price = payload.unit_price
    rec.total_price = payload.total_price
    rec.fuel_type = payload.fuel_type
    rec.note = payload.note
    db.commit()
    db.refresh(rec)
    all_records = _query_records(db, payload.vehicle_id)
    enriched = enrich_records(all_records)
    for e in enriched:
        if e["id"] == rec.id:
            return e
    return rec

@app.delete("/api/records/{record_id}")
def delete_record(record_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rec = db.query(FuelRecord).filter(FuelRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="记录不存在")
    if rec.vehicle_id:
        _get_user_vehicle(db, user, rec.vehicle_id)
    db.delete(rec)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Stats (auth required)
# ---------------------------------------------------------------------------
@app.get("/api/stats", response_model=StatsOut)
def get_stats(vehicle_id: int = Query(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _get_user_vehicle(db, user, vehicle_id)
    records = _query_records(db, vehicle_id)
    count = len(records)
    if count == 0:
        return StatsOut(total_mileage=0, total_cost=0, total_volume=0, avg_consumption=None, daily_mileage=None, record_count=0)
    first, last = records[0], records[-1]
    total_mileage = last.mileage
    driven_mileage = last.mileage - first.mileage if count > 1 else 0
    total_cost = sum(r.total_price or 0 for r in records)
    total_volume_all = sum(r.volume or 0 for r in records)
    total_volume = sum(r.volume or 0 for r in records[1:]) if count > 1 else 0
    avg_consumption = round(total_volume / driven_mileage * 100, 2) if driven_mileage > 0 and total_volume > 0 else None
    days_span = (last.date - first.date).days if count > 1 else 0
    daily_mileage = round(driven_mileage / days_span, 1) if days_span > 0 else None
    return StatsOut(total_mileage=round(total_mileage, 1), total_cost=round(total_cost, 2),
                    total_volume=round(total_volume_all, 2), avg_consumption=avg_consumption,
                    daily_mileage=daily_mileage, record_count=count)


# ---------------------------------------------------------------------------
# Period summary (auth required)
# ---------------------------------------------------------------------------
@app.get("/api/stats/summary", response_model=PeriodSummaryOut)
def get_period_summary(
    vehicle_id: int = Query(...),
    mode: str = Query("yearly", pattern="^(yearly|monthly)$"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_user_vehicle(db, user, vehicle_id)
    all_records = _query_records(db, vehicle_id)
    if not all_records:
        return PeriodSummaryOut(mode=mode, items=[])
    enriched: list[dict] = []
    for i, rec in enumerate(all_records):
        dist = 0.0
        if i > 0:
            dist = rec.mileage - all_records[i - 1].mileage
            if dist < 0: dist = 0.0
        enriched.append({"date": rec.date, "volume": rec.volume or 0, "total_price": rec.total_price or 0, "distance": dist, "is_first": i == 0})
    def period_key(d: date) -> str:
        return str(d.year) if mode == "yearly" else f"{d.year}-{d.month:02d}"
    groups: dict[str, list[dict]] = OrderedDict()
    for e in enriched:
        groups.setdefault(period_key(e["date"]), []).append(e)
    items: list[PeriodSummaryItem] = []
    for period, recs in groups.items():
        total_dist = sum(r["distance"] for r in recs)
        total_vol = sum(r["volume"] for r in recs if not r["is_first"])
        total_cost = sum(r["total_price"] for r in recs)
        cnt = len(recs)
        avg_cons = round(total_vol / total_dist * 100, 2) if total_dist > 0 and total_vol > 0 else None
        dates = [r["date"] for r in recs]
        day_span = (max(dates) - min(dates)).days if len(dates) > 1 else 0
        daily_mi = round(total_dist / day_span, 1) if day_span > 0 else None
        items.append(PeriodSummaryItem(period=period, mileage=round(total_dist, 1), volume=round(total_vol, 2),
                                       cost=round(total_cost, 2), avg_consumption=avg_cons, daily_mileage=daily_mi, record_count=cnt))
    items.reverse()
    return PeriodSummaryOut(mode=mode, items=items)


# ---------------------------------------------------------------------------
# CSV export (auth required)
# ---------------------------------------------------------------------------
@app.get("/api/export/csv")
def export_csv(vehicle_id: int = Query(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _get_user_vehicle(db, user, vehicle_id)
    records = _query_records(db, vehicle_id)
    enriched = enrich_records(records)
    enriched.reverse()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["日期", "里程数(km)", "油量(L)", "单价(元/L)", "总价(元)", "油号", "区间里程(km)", "油耗(L/100km)", "每公里费用(元/km)", "备注"])
    for r in enriched:
        writer.writerow([r["date"], r["mileage"], r["volume"] or "", r["unit_price"] or "",
                         r["total_price"] or "", r["fuel_type"] or "", r["distance"] or "",
                         r["fuel_consumption"] or "", r["cost_per_km"] or "", r["note"] or ""])
    output.seek(0)
    content = "﻿" + output.getvalue()
    return StreamingResponse(io.BytesIO(content.encode("utf-8-sig")), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=fuel_records.csv"})


# ---------------------------------------------------------------------------
# Import helpers (shared by xlsx / csv)
# ---------------------------------------------------------------------------
_HEADER_KEYWORDS = {
    "date": ["加油日期", "日期"],
    "volume": ["加油量", "油量"],
    "unit_price": ["支付单价", "单价"],
    "total_price": ["支付总额", "总价", "总额"],
    "mileage": ["行驶里程", "里程数", "里程"],
    "fuel_type": ["油号", "燃油类型"],
    "note": ["备注"],
}

_SKIP_KEYWORDS = ["总记录", "总增加", "总加油", "总支付", "平均油耗", "统计汇总", "用户ID", "weCarId"]


def _build_col_map(header_row: list[str]) -> dict[str, int]:
    """Map logical field names to column indices by matching Chinese header keywords."""
    col_map: dict[str, int] = {}
    for idx, h in enumerate(header_row):
        for field, keywords in _HEADER_KEYWORDS.items():
            if field not in col_map:
                for kw in keywords:
                    if kw in h:
                        col_map[field] = idx
                        break
    return col_map


def _import_rows(
    rows: list[tuple | list],
    col_map: dict[str, int],
    vehicle_id: int,
    db: Session,
    row_offset: int = 2,
) -> tuple[int, int, list[str]]:
    """Process data rows and insert FuelRecords. Returns (imported, skipped, errors)."""
    existing: set[tuple[str, float]] = set()
    for rec in db.query(FuelRecord.date, FuelRecord.mileage).filter(FuelRecord.vehicle_id == vehicle_id).all():
        existing.add((str(rec.date), float(rec.mileage)))

    imported = 0; skipped = 0; errors: list[str] = []
    for i, row in enumerate(rows):
        row_idx = i + row_offset
        first_cell = row[0] if row else None
        if first_cell is None:
            continue
        if isinstance(first_cell, str) and any(kw in first_cell for kw in _SKIP_KEYWORDS):
            continue
        try:
            # --- date ---
            raw_date = row[col_map["date"]] if "date" in col_map else None
            if raw_date is None:
                continue
            if isinstance(raw_date, datetime):
                rec_date = raw_date.date()
            elif isinstance(raw_date, date):
                rec_date = raw_date
            elif isinstance(raw_date, str):
                raw_date = raw_date.strip()
                if not raw_date:
                    continue
                rec_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
            else:
                errors.append(f"第 {row_idx} 行: 无法解析日期 '{raw_date}'")
                continue

            # --- mileage ---
            raw_mileage = row[col_map["mileage"]] if "mileage" in col_map else None
            if raw_mileage is None:
                continue
            mileage = float(raw_mileage)
            if (str(rec_date), mileage) in existing:
                skipped += 1
                continue

            # --- optional fields ---
            def _float(field: str):
                if field not in col_map:
                    return None
                val = row[col_map[field]] if col_map[field] < len(row) else None
                if val is None or (isinstance(val, str) and not val.strip()):
                    return None
                return float(val)

            volume = _float("volume")
            unit_price = _float("unit_price")
            total_price = _float("total_price")

            fuel_type = None
            if "fuel_type" in col_map and col_map["fuel_type"] < len(row) and row[col_map["fuel_type"]] is not None:
                ft = str(row[col_map["fuel_type"]]).strip()
                if ft:
                    fuel_type = ft

            note_val = ""
            if "note" in col_map and col_map["note"] < len(row) and row[col_map["note"]] is not None:
                note_val = str(row[col_map["note"]]).strip()

            record = FuelRecord(
                vehicle_id=vehicle_id, date=rec_date, mileage=mileage,
                volume=volume, unit_price=unit_price, total_price=total_price,
                fuel_type=fuel_type, note=note_val,
            )
            db.add(record)
            existing.add((str(rec_date), mileage))
            imported += 1
        except Exception as e:
            errors.append(f"第 {row_idx} 行: {e}")

    if imported > 0:
        db.commit()
    return imported, skipped, errors


def _parse_csv_bytes(contents: bytes) -> tuple[list[str], list[list[str]]]:
    """Decode CSV bytes (auto-detect encoding) and return (header, data_rows)."""
    text = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = contents.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        raise ValueError("无法识别 CSV 文件编码，请使用 UTF-8 或 GBK 编码")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if len(all_rows) < 1:
        raise ValueError("CSV 文件为空")
    header = [h.strip() for h in all_rows[0]]
    data = [row for row in all_rows[1:] if any(cell.strip() for cell in row)]
    return header, data


# ---------------------------------------------------------------------------
# Import file — unified endpoint (xlsx / csv)
# ---------------------------------------------------------------------------
@app.post("/api/import/file", response_model=ImportResult)
async def import_file(
    file: UploadFile = File(...), vehicle_id: int = Form(...),
    user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    _get_user_vehicle(db, user, vehicle_id)
    filename = (file.filename or "").lower()
    contents = await file.read()

    if filename.endswith((".xlsx", ".xls")):
        # --- Excel ---
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件: {e}")
        ws = wb[wb.sheetnames[0]]
        header_row = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    elif filename.endswith(".csv"):
        # --- CSV ---
        try:
            header_row, data_rows = _parse_csv_bytes(contents)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    else:
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .csv 格式的文件")

    col_map = _build_col_map(header_row)
    required = ["date", "mileage"]
    missing = [f for f in required if f not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"文件中缺少必要的列: {', '.join(missing)}。找到的表头: {header_row}")

    imported, skipped, errors = _import_rows(data_rows, col_map, vehicle_id, db)
    return ImportResult(imported=imported, skipped=skipped, errors=errors[:20])


# Keep old path for backwards compatibility
@app.post("/api/import/xlsx", response_model=ImportResult)
async def import_xlsx_compat(
    file: UploadFile = File(...), vehicle_id: int = Form(...),
    user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    return await import_file(file=file, vehicle_id=vehicle_id, user=user, db=db)


# ---------------------------------------------------------------------------
# Static files (serve frontend)
# ---------------------------------------------------------------------------
app.mount("/", StaticFiles(directory="static", html=True), name="static")
